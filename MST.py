
from asyncio import QueueEmpty
from cProfile import label
from errno import EDEADLK
from queue import Queue
from random import randint
import time
import copy
from unicodedata import name
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import xlrd

from numpy import array, size
plt.rcParams["font.sans-serif"] = ["SimHei"]  # 设置字体
plt.rcParams["axes.unicode_minus"] = False  # 该语句解决图像中的“-”负号的乱码问题
# 随机生成图

NoCount = 10000000000
nameArray = ["Krushal", "Prim", "Prim_Array", "Sollin"]
ConnectArray = []
# 定义写入的Excel文件
wb = Workbook()

ws = wb.create_sheet("che")
ws.append(["时间"]+nameArray)


def makeMST(size, value):
    MST = [[0] * size for i in range(size)]
    for i in range(size):
        for j in range(i):
            MST[i][j] = randint(0, value)
            MST[j][i] = MST[i][j]
    return MST

#  输出图


def PrintMST(MST):
    print("生成的无向图为")
    for ele in MST:
        print(ele)

#  Krushal 算法生成MST


def Krushal(MST):
    time_start = time.time()
    size = len(MST)
    # 用来统计不相交集合的数组
    JointSet = [[i] for i in range(size)]
    # 用来计算Edge的数组
    EdgeArray = []
    # 最小生成树的数组
    MSTArray = []
    # 将所有连通的边放到数组中 进行排序

    # 上一次检测到的边
    last_index = 0
    for i in range(size):
        for j in range(i):
            if MST[i][j] != 0:
                EdgeArray.append([i, j, MST[i][j]])
    EdgeArray = sorted(EdgeArray, key=lambda book: book[2])
    # print("当前的边排序为", EdgeArray)
    # 进行贪心生成MST
    for i in range(size-1):
        for index in range(last_index, len(EdgeArray)):
            ele = EdgeArray[index]
            if JointSetJudge(JointSet, ele[0], ele[1]):
                last_index = index
                JointSetAppend(JointSet, ele[0], ele[1])
                MSTArray.append(EdgeArray[index])
                del EdgeArray[index]
                break
            if(index == len(EdgeArray)-1):
                print("当前图可能不连通，请重试！")
    # print("Krushal MST为", MSTArray)
    # print("Krushal MST SUM 为", sum([MSTArray[i][2]
    #       for i in range(len(MSTArray))]))
    time_end = time.time()
    # print("Krushal 算法程序执行时间为s", time_end-time_start, "s")
    return time_end-time_start
#  判断一条边是否能进行添加

#  Prim 算法生成 MST


def Prim(MST):
    time_start = time.time()
    size = len(MST)
    #  所有边的数组
    EdgeArray = []
    # MST 的数组
    MSTArray = []
    # 当前已经选取的节点
    YesNodeArray = [len(MST)-1]
    # 当前尚未选取的节点
    NoNodeArray = [i for i in range(len(MST)-1)]
    # 将所有连通的边 放到数组中
    for i in range(size):
        for j in range(i):
            if MST[i][j] != 0:
                EdgeArray.append([i, j, MST[i][j]])
    EdgeArray = sorted(EdgeArray, key=lambda book: book[2])
    #  进行贪心算法 生成MST
    last_index = 0
    for _ in range(len(MST) - 1):
        for index in range(0, len(EdgeArray)):
            # print(last_index)
            if (EdgeArray[index][0] in YesNodeArray and EdgeArray[index][1] in NoNodeArray):
                NoNodeArray.remove(EdgeArray[index][1])
                YesNodeArray.append(EdgeArray[index][1])
                MSTArray.append(EdgeArray[index])
                del EdgeArray[index]
                break
            if (EdgeArray[index][1] in YesNodeArray and EdgeArray[index][0] in NoNodeArray):
                NoNodeArray.remove(EdgeArray[index][0])
                YesNodeArray.append(EdgeArray[index][0])
                MSTArray.append(EdgeArray[index])
                del EdgeArray[index]
                break
        if(index == len(EdgeArray)-1):
            print("当前图可能不连通，请重试！")
    # print("Prim MST为", MSTArray)
    # print("Prim MST SUM 为", sum([MSTArray[i][2]
        #   for i in range(len(MSTArray))]))
    time_end = time.time()
    # print("Prim 程序执行时间为s", time_end-time_start, "s")
    return time_end-time_start


def Prim_Array(MST):
    time_start = time.time()
    size = len(MST)
    # MST 的数组
    ConnectArray = [0 for i in range(len(MST))]
    MSTArray = []
    MSTSUM = 0
    # 将所有连通的边 放到数组中
    EdgeArray = MST[len(MST)-1]
    # 记得对第一个点进行初始化
    EdgeArray[len(MST)-1] = NoCount
    #  进行贪心算法 生成MST
    last_index = 0

    for i in range(len(MST)-1):
        # 返回当前已经连通的最小值
        ZeroIndex = ReturnMinIndex(EdgeArray)
        # print("---", EdgeArray,  MSTSUM, ZeroIndex)
        MSTSUM += EdgeArray[ZeroIndex]
        EdgeArray = MergeArray(EdgeArray, MST[ZeroIndex], ZeroIndex)
    # print("Prim_Array MST SUM为", MSTSUM)
    time_end = time.time()
    # print("Prim_Array 程序执行时间为s", time_end-time_start, "s")
    return time_end-time_start


def ReturnMinIndex(array):
    Min = 10000000000
    MinIndex = 100000000
    for index in range(len(array)):
        if array[index] != NoCount and array[index] != 0 and array[index] < Min:
            MinIndex = index
            Min = array[index]
    return MinIndex


def MergeArray(array1, array2, zero):
    # print("进来的结果", array1, array2, zero)
    for index in range(len(array1)):
        if array1[index] == NoCount or array2[index] == NoCount:
            array1[index] == NoCount
            continue
        if array1[index] != 0 and array2[index] != 0:
            if array1[index] > array2[index]:
                array1[index] = array2[index]
            # ConnectArray[index] = zero
            continue
        if array1[index] == 0 and array2[index] != 0:
            array1[index] = array2[index]
            # ConnectArray[index] = zero
            continue
    array1[zero] = NoCount
    # print("出去的结果", array1)
    return array1


def Sollin(MST):
    time_start = time.time()
    size = len(MST)
    #  所有边的数组
    EdgeArray = []
    #  MST 数组
    MSTArray = []
    for i in range(size):
        for j in range(i):
            if MST[i][j] != 0:
                EdgeArray.append([i, j, MST[i][j]])
    EdgeArray = sorted(EdgeArray, key=lambda book: book[2])
    QueueNodeArray = [[i] for i in range(size)]
    while len(QueueNodeArray) != 1:
        set = QueueNodeArray.pop(0)
        edge = GetConnectMinumumEdge(set, EdgeArray)
        EdgeArray.remove(edge)
        MSTArray.append(edge)
        index1 = edge[0]
        index2 = edge[1]
        # print(QueueNodeArray,  set, "\n",  edge)
        if index1 in set:
            NodeArrayIndex = GetHaveIndexElement(index2, QueueNodeArray)
            # print("与之相连的点索引为", NodeArrayIndex)
            QueueNodeArray.append(set + QueueNodeArray[NodeArrayIndex])
            QueueNodeArray.remove(QueueNodeArray[NodeArrayIndex])
        if index2 in set:
            NodeArrayIndex = GetHaveIndexElement(index1, QueueNodeArray)
            # print("与之相连的点索引为", NodeArrayIndex)
            QueueNodeArray.append(set + QueueNodeArray[NodeArrayIndex])
            QueueNodeArray.remove(QueueNodeArray[NodeArrayIndex])
    # print("Sollin MST为", MSTArray)
    # print("Sollin MST SUM 为", sum([MSTArray[i][2]
        #   for i in range(len(MSTArray))]))
    time_end = time.time()
    # print("Sollin 程序执行时间为s", time_end-time_start, "s")
    return time_end-time_start


def GetHaveIndexElement(point, Array):
    for index in range(len(Array)):
        if point in Array[index]:
            return index
    return None


def GetConnectMinumumEdge(set, EdgeArray):
    for index in range(len(EdgeArray)):
        ele = EdgeArray[index]
        index1 = ele[0]
        index2 = ele[1]
        if (index1 in set and index2 not in set) or (index1 not in set and index2 in set):
            return EdgeArray[index]
    return None


def JointSetJudge(JointSet, value1, value2):
    value1Index = JointSetHave(JointSet, value1)
    value2Index = JointSetHave(JointSet, value2)
    return value1Index != value2Index

#  不相交集 append


def JointSetAppend(JointSet, value1, value2):
    value1Index = JointSetHave(JointSet, value1)
    value2Index = JointSetHave(JointSet, value2)
    # print("不相交集合为", JointSet, value1Index, value2Index)
    JointSet[value1Index] += (JointSet[value2Index])
    del JointSet[value2Index]
    pass

#  判断是哪个不相交集拥有值


def JointSetHave(JointSet, value):
    for i in range(len(JointSet)):
        if(value in JointSet[i]):
            return i
    return None


def WriteExpdata(KTime, PrimT, PrimA_T, Sollin_T, Size):
    f = open("experiment_data.txt", "a+")
    string = "Krushal Time is %6f , Prim Time is %6f , Prim_Array Time is %6f , Sollin Time is %6f Size is %d \n" % (
        KTime, PrimT, PrimA_T, Sollin_T, Size)
    f.write(string)


def WriteExpdata(KTime, PrimT, PrimA_T, Sollin_T, Size, Max):
    f = open("experiment_data_"+str(Max)+".txt", "a+")
    string = "Krushal Time is %6f , Prim Time is %6f , Prim_Array Time is %6f , Sollin Time is %6f Size is %d \n" % (
        KTime, PrimT, PrimA_T, Sollin_T, Size)
    f.write(string)
#  用来做调试以及可视化数据的函数


def WriteIntoExcel(col):
    ws.append(col)


def main():
    print("输入图的大小")
    size = int(input())
    print("输入图的边的最大值")
    value = int(input())
    MST = makeMST(size, value)
    # print(MST)
    KrushalTime = Krushal(MST)
    PrimTime = Prim(MST)
    Prim_Array_Time = Prim_Array(copy.deepcopy(MST))
    # Prim_Array()
    SollinTime = Sollin(MST)

    pass


def DrawPicture(x, y, nameArray):
    fig, ax = plt.subplots()  # 创建图实例
    # x = np.linspace(0, 2, 100)  # 创建x的取值范围
    for index in range(len(y)):
        ax.plot(x, y[index], label=nameArray[index])
    ax.set_xlabel('数据大小')  # 设置x轴名称 x label
    ax.set_ylabel('算法时间')  # 设置y轴名称 y label
    ax.set_title('算法运行时间展示')  # 设置图名为Simple Plot
    ax.legend()  # 自动检测要在图例中显示的元素，并且显示
    plt.show()


def main2():
    print("绘制图表的范围")
    End = int(input())
    Begin = 10
    x = []
    y = [[], [], [], []]

    while(Begin <= End):
        MST = makeMST(Begin, Begin)
        KrushalTime = Krushal(MST)
        PrimTime = Prim(MST)
        Prim_Array_Time = Prim_Array(copy.deepcopy(MST))
        # Prim_Array()
        SollinTime = Sollin(MST)
        x.append(Begin)
        y[0].append(KrushalTime)
        y[1].append(PrimTime)
        y[2].append(Prim_Array_Time)
        y[3].append(SollinTime)
        WriteExpdata(KrushalTime, PrimTime, Prim_Array_Time,
                     SollinTime, Begin, End)
        WriteIntoExcel([Begin, KrushalTime*1000, PrimTime*1000,
                       Prim_Array_Time*1000, SollinTime*1000])
        Begin += 10
    # print(x, y, nameArray)
    wb.save("experimentData_"+str(End)+".xlsx")
    DrawPicture(x, y, nameArray)


def main3():
    print("请输入你所要读取的文件")
    filename = input()
    print(filename)
    wb2 = load_workbook(filename.strip("\n"))
    che = wb2["che"]
    x = []
    y = [[], [], [], []]
    firstRow = True
    for ele in che.values:
        if firstRow:
            firstRow = False
            continue
        # print(ele)
        x.append(ele[0])
        y[0].append(ele[1])
        y[1].append(ele[2])
        y[2].append(ele[3])
        y[3].append(ele[4])
    # print(x, y, nameArray
    DrawPicture(x, y, nameArray)


main3()
